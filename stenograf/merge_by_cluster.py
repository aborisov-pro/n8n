#!/usr/bin/env python3
"""
Склейка лекций по кластерам из xlsx-маппинга.

Читает «Распределение файлов.xlsx» (колонки: Файл | Последовательность | Кластер),
склеивает файлы каждого кластера по возрастанию «Последовательности» в единый MP3,
оптимальный для транскрибации (mono 16 kHz, 64 kbps), со срезом тишины и loudnorm.

Зависимости:
    pip install openpyxl
    ffmpeg в PATH

Запуск (из папки «записи лекций»):
    python merge_by_cluster.py

Или с явными путями:
    python merge_by_cluster.py --folder "C:\\...\\записи лекций" ^
        --mapping "C:\\...\\Распределение файлов.xlsx" --out "C:\\...\\merged"

Флаги:
    --no-silence-remove   НЕ срезать тишину
    --no-loudnorm         НЕ нормализовать громкость
    --bitrate 64k         битрейт MP3 (по умолчанию 64k)
    --sample-rate 16000   частота дискретизации (по умолчанию 16000)
    --jobs N              параллельные pre-encode задачи (по умолчанию 4)
    --dry-run             только показать план, ничего не делать
"""
from __future__ import annotations

import argparse
import concurrent.futures
import glob
import os
import re
import subprocess
import sys
import tempfile
import time
from collections import defaultdict
from pathlib import Path

# Максимальная длительность одного выходного файла в секундах.
# 13800 с = 3ч 50м — запас до лимита 4ч у сервисов транскрибации.
MAX_PART_SEC = 13800

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Нужен openpyxl. Установи:  pip install openpyxl")


# Транслитерация ru→lat для имён выходных файлов
CYR_LAT = {
    "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "yo",
    "ж": "zh", "з": "z", "и": "i", "й": "y", "к": "k", "л": "l", "м": "m",
    "н": "n", "о": "o", "п": "p", "р": "r", "с": "s", "т": "t", "у": "u",
    "ф": "f", "х": "h", "ц": "c", "ч": "ch", "ш": "sh", "щ": "sch", "ъ": "",
    "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya",
}

# Подстановки перед посимвольной транслитерацией — для читаемости
CLUSTER_SUBS = [
    ("Пауэр Квери", "Power_Query"),
    ("Пауэр квери", "Power_Query"),
    ("Power Query", "Power_Query"),
    ("вебинар", "webinar"),
    ("н8н", "n8n"),
    ("марта", "Marta"),
    ("апреля", "Aprelya"),
    ("мая", "Maya"),
    ("июня", "Iyunya"),
]


def transliterate(s: str) -> str:
    for old, new in CLUSTER_SUBS:
        s = s.replace(old, new)
    out = []
    for ch in s:
        low = ch.lower()
        if low in CYR_LAT:
            tr = CYR_LAT[low]
            out.append(tr.capitalize() if ch.isupper() and tr else tr)
        else:
            out.append(ch)
    s = "".join(out)
    s = re.sub(r"\s+", "_", s.strip())
    s = re.sub(r"[^\w\-.]", "", s)
    s = re.sub(r"_+", "_", s)
    return s


def run(cmd: list[str]) -> None:
    r = subprocess.run(cmd, capture_output=True, text=True)
    if r.returncode != 0:
        sys.stderr.write(f"\nFFMPEG ERR:\n{r.stderr[-2000:]}\n")
        raise RuntimeError(f"ffmpeg rc={r.returncode}")


def probe_duration(path: str) -> float:
    try:
        r = subprocess.run(
            ["ffprobe", "-v", "error", "-show_entries", "format=duration",
             "-of", "default=nk=1:nw=1", path],
            capture_output=True, text=True, check=True,
        )
        return float(r.stdout.strip())
    except Exception:
        return 0.0


def read_mapping(xlsx_path: str) -> dict[str, list[tuple[int, str]]]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    try:
        i_file = header.index("файл")
        i_seq = header.index("последовательность")
        i_cluster = header.index("кластер")
    except ValueError:
        sys.exit("В xlsx ожидаются колонки: Файл | Последовательность | Кластер")
    groups: dict[str, list[tuple[int, str]]] = defaultdict(list)
    for row in rows[1:]:
        if not row or not row[i_file]:
            continue
        fname = str(row[i_file]).strip()
        seq = int(row[i_seq]) if row[i_seq] is not None else 0
        cluster = str(row[i_cluster]).strip() if row[i_cluster] else "_uncategorized"
        groups[cluster].append((seq, fname))
    for k in groups:
        groups[k].sort()
    return groups


def preencode(src: str, dst: str, sr: str, br: str) -> None:
    run([
        "ffmpeg", "-hide_banner", "-loglevel", "error", "-y",
        "-i", src, "-ac", "1", "-ar", sr,
        "-c:a", "libmp3lame", "-b:a", br, dst,
    ])


def _build_filters(silence_remove: bool, loudnorm: bool) -> list[str]:
    filters = []
    if silence_remove:
        filters.append(
            "silenceremove=stop_periods=-1:stop_duration=2:stop_threshold=-40dB"
        )
    if loudnorm:
        filters.append("loudnorm=I=-16:TP=-1.5:LRA=11")
    return filters


def _existing_outputs(out_dir: str, basename: str) -> list[str]:
    """Собирает уже существующие выходы: одиночный или части _part1, _part2, ..."""
    single = os.path.join(out_dir, f"{basename}.mp3")
    hits = [single] if os.path.exists(single) else []
    part_rx = re.compile(rf"^{re.escape(basename)}_part\d+\.mp3$")
    for fn in os.listdir(out_dir):
        if part_rx.match(fn):
            hits.append(os.path.join(out_dir, fn))
    return sorted(hits)


def _finalize_segments(out_dir: str, basename: str, seg_prefix: str) -> list[str]:
    """Собирает сегменты seg_prefix_N.mp3 и переименовывает в итоговые имена."""
    seg_rx = re.compile(rf"^{re.escape(seg_prefix)}(\d+)\.mp3$")
    segs = []
    for fn in os.listdir(out_dir):
        m = seg_rx.match(fn)
        if m:
            segs.append((int(m.group(1)), os.path.join(out_dir, fn)))
    segs.sort()
    results = []
    if len(segs) == 1:
        dst = os.path.join(out_dir, f"{basename}.mp3")
        os.rename(segs[0][1], dst)
        results.append(dst)
    else:
        for i, (_, src) in enumerate(segs, 1):
            dst = os.path.join(out_dir, f"{basename}_part{i}.mp3")
            os.rename(src, dst)
            results.append(dst)
    return results


def merge_cluster(
    cluster: str, items: list[tuple[int, str]], folder: str, out_dir: str,
    sr: str, br: str, loudnorm: bool, silence_remove: bool, jobs: int,
) -> None:
    basename = transliterate(cluster)
    existing = _existing_outputs(out_dir, basename)
    if existing:
        names = [os.path.basename(p) for p in existing]
        print(f"  [skip] уже существует: {names}")
        return

    srcs = []
    missing = []
    for seq, fname in items:
        p = os.path.join(folder, fname)
        if os.path.exists(p):
            srcs.append((seq, p))
        else:
            missing.append(fname)
    if missing:
        print(f"  [WARN] пропущены (нет на диске): {missing}")
    if not srcs:
        print(f"  [skip] файлов не найдено")
        return

    print(f"  {len(srcs)} файл(ов) → {basename}[_partN].mp3")

    filters = _build_filters(silence_remove, loudnorm)
    seg_prefix = f"{basename}__segtmp_"
    seg_pattern = os.path.join(out_dir, f"{seg_prefix}%d.mp3")

    # Общий хвост команды: segment-muxer вместо одиночного выхода
    tail = [
        "-ac", "1", "-ar", sr,
        "-c:a", "libmp3lame", "-b:a", br,
        "-f", "segment",
        "-segment_time", str(MAX_PART_SEC),
        "-reset_timestamps", "1",
        seg_pattern,
    ]

    if len(srcs) == 1:
        # Один файл: без concat
        _, src = srcs[0]
        cmd = ["ffmpeg", "-hide_banner", "-loglevel", "warning", "-y", "-i", src]
        if filters:
            cmd += ["-af", ",".join(filters)]
        cmd += tail
        run(cmd)
    else:
        with tempfile.TemporaryDirectory(prefix="mrg_") as tmp:
            # Параллельный pre-encode в одинаковый формат
            tmps = [os.path.join(tmp, f"{i:03d}.mp3") for i in range(len(srcs))]
            with concurrent.futures.ThreadPoolExecutor(max_workers=jobs) as ex:
                futs = [ex.submit(preencode, src, dst, sr, br)
                        for (_, src), dst in zip(srcs, tmps)]
                for f in futs:
                    f.result()
            # Список для concat-демуксера
            lst = os.path.join(tmp, "list.txt")
            with open(lst, "w", encoding="utf-8") as fh:
                for p in tmps:
                    esc = p.replace("'", "'\\''")
                    fh.write(f"file '{esc}'\n")
            # Финал: concat → [silenceremove] → [loudnorm] → segment muxer
            cmd = [
                "ffmpeg", "-hide_banner", "-loglevel", "warning", "-y",
                "-f", "concat", "-safe", "0", "-i", lst,
            ]
            if filters:
                cmd += ["-af", ",".join(filters)]
            cmd += tail
            run(cmd)

    finals = _finalize_segments(out_dir, basename, seg_prefix)
    for p in finals:
        sz = os.path.getsize(p) / 1024 / 1024
        dur = probe_duration(p)
        print(f"  готово: {os.path.basename(p)}  ({sz:.1f} МБ, {dur/3600:.2f} ч)")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--folder", default=".", help="Папка с источниками")
    ap.add_argument("--mapping", default=None, help="Путь к xlsx (по умолчанию: Распределение файлов.xlsx в --folder)")
    ap.add_argument("--out", default=None, help="Папка вывода (по умолчанию: <folder>/merged)")
    ap.add_argument("--silence-remove", action="store_true", default=True)
    ap.add_argument("--no-silence-remove", dest="silence_remove", action="store_false")
    ap.add_argument("--loudnorm", action="store_true", default=True)
    ap.add_argument("--no-loudnorm", dest="loudnorm", action="store_false")
    ap.add_argument("--bitrate", default="64k")
    ap.add_argument("--sample-rate", default="16000")
    ap.add_argument("--jobs", type=int, default=4)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    folder = os.path.abspath(args.folder)
    if not os.path.isdir(folder):
        sys.exit(f"Нет папки: {folder}")

    mapping = args.mapping or os.path.join(folder, "Распределение файлов.xlsx")
    if not os.path.isfile(mapping):
        sys.exit(f"Нет xlsx-маппинга: {mapping}")

    out_dir = args.out or os.path.join(folder, "merged")
    os.makedirs(out_dir, exist_ok=True)

    groups = read_mapping(mapping)
    total_files = sum(len(v) for v in groups.values())
    print(f"Кластеров: {len(groups)}, файлов в маппинге: {total_files}")
    print(f"Источник: {folder}")
    print(f"Вывод:    {out_dir}")
    print(f"Параметры: sr={args.sample_rate}, br={args.bitrate}, "
          f"silence_remove={args.silence_remove}, loudnorm={args.loudnorm}, jobs={args.jobs}")
    print()

    if args.dry_run:
        for cluster, items in groups.items():
            print(f"[{cluster}] → {transliterate(cluster)}.mp3")
            for seq, fname in items:
                print(f"   {seq:>3}. {fname}")
        return

    t0 = time.time()
    for cluster, items in groups.items():
        print(f"[{cluster}]")
        t_cl = time.time()
        try:
            merge_cluster(
                cluster, items, folder, out_dir,
                args.sample_rate, args.bitrate,
                args.loudnorm, args.silence_remove, args.jobs,
            )
        except Exception as e:
            print(f"  ОШИБКА: {e}")
        print(f"  время: {(time.time()-t_cl)/60:.1f} мин\n")
    print(f"Всего: {(time.time()-t0)/60:.1f} мин")


if __name__ == "__main__":
    main()
